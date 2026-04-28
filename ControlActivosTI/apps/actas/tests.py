from datetime import date
from pathlib import Path
import shutil
import uuid
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.core.files.storage import default_storage
from django.test import TestCase
from django.test.utils import override_settings
from django.utils import timezone
from openpyxl import load_workbook

from apps.actas.models import ActaEntrega
from apps.actas.services import generar_o_actualizar_acta
from apps.activos.models import Activo
from apps.asignaciones.models import Asignacion, AsignacionDetalle
from apps.catalogos.models import Area, Cargo, CentroCosto, EstadoActivo, TipoActivo, Ubicacion
from apps.colaboradores.models import Colaborador


def make_test_media_root():
    media_root = Path.cwd() / "test-media" / uuid.uuid4().hex
    media_root.mkdir(parents=True, exist_ok=True)
    return media_root


class ActaEntregaExcelTests(TestCase):
    def setUp(self):
        self.media_root = make_test_media_root()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()

        self.user = get_user_model().objects.create_user(
            username="responsable",
            password="secret123",
            first_name="Juan",
            last_name="Villacres",
        )
        self.area = Area.objects.create(nombre="TI")
        self.cargo = Cargo.objects.create(nombre="Analista de soporte")
        self.ubicacion = Ubicacion.objects.create(nombre="Matriz")
        self.ceco = CentroCosto.objects.create(
            codigo="TI001",
            nombre="Tecnologia",
            acepta_asignaciones=True,
            activo=True,
        )
        self.tipo_activo = TipoActivo.objects.create(nombre="Laptop")
        self.estado_disponible = EstadoActivo.objects.create(
            nombre="Disponible",
            permite_asignacion=True,
        )
        EstadoActivo.objects.create(nombre="Asignado", permite_asignacion=False)
        self.colaborador = Colaborador.objects.create(
            nombres="Ana",
            apellidos="Perez",
            cedula="0123456789",
            correo_corporativo="ana.perez@example.com",
            cargo=self.cargo,
            area=self.area,
            ubicacion=self.ubicacion,
            centro_costo=self.ceco,
            fecha_ingreso=date(2024, 1, 10),
        )
        self.asignacion = Asignacion.objects.create(
            colaborador=self.colaborador,
            fecha_asignacion=date(2026, 4, 20),
            usuario_responsable=self.user,
        )

    def tearDown(self):
        self.override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def crear_detalle(self, orden=1, **kwargs):
        activo = Activo.objects.create(
            tipo_activo=self.tipo_activo,
            marca=kwargs.get("marca", "Dell"),
            modelo=kwargs.get("modelo", f"Latitude {orden}"),
            serie=kwargs.get("serie", f"SERIE{orden:03d}"),
            cpu=kwargs.get("cpu", "Intel i5"),
            ram=kwargs.get("ram", "16GB"),
            disco=kwargs.get("disco", "512GB SSD"),
            sistema_operativo=kwargs.get("sistema_operativo", "Windows 11"),
            valor=kwargs.get("valor", "1200.50"),
            estado_activo=self.estado_disponible,
        )
        return AsignacionDetalle.objects.create(
            asignacion=self.asignacion,
            activo=activo,
            orden=orden,
        )

    def cargar_workbook_generado(self):
        acta = generar_o_actualizar_acta(self.asignacion, self.user)
        with default_storage.open(acta.archivo.name, "rb") as archivo:
            workbook = load_workbook(archivo)
        return acta, workbook

    def cargar_xml_hoja_generada(self, acta):
        with default_storage.open(acta.archivo.name, "rb") as archivo:
            with ZipFile(archivo) as paquete:
                return paquete.read("xl/worksheets/sheet1.xml")

    def test_generates_delivery_excel_from_template_with_assignment_data(self):
        self.crear_detalle()

        acta, workbook = self.cargar_workbook_generado()
        ws = workbook.active

        self.assertEqual(acta.tipo, ActaEntrega.TipoActa.ENTREGA)
        self.assertTrue(acta.nombre_archivo.endswith(".xlsx"))
        self.assertEqual(ws["E6"].value.date(), timezone.localdate())
        self.assertEqual(ws["E10"].value, "Ana Perez")
        self.assertEqual(ws["E11"].value, "0123456789")
        self.assertEqual(ws["I10"].value, "Analista de soporte")
        self.assertIsNone(ws["B1"].value)
        self.assertEqual(ws["B14"].value, "Laptop")
        self.assertEqual(ws["D14"].value, "Dell")
        self.assertEqual(ws["F14"].value, 1200.50)
        self.assertEqual(ws["F14"].number_format, "$#,##0.00")
        self.assertIsNone(ws["G14"].value)
        self.assertIn("CPU: Intel i5", ws["H14"].value)
        self.assertIn("RAM: 16GB", ws["H14"].value)
        self.assertTrue(ws["H14"].alignment.wrap_text)
        self.assertGreater(ws.row_dimensions[14].height, 30)
        self.assertIsNone(ws["I14"].value)
        xml = self.cargar_xml_hoja_generada(acta)
        self.assertNotIn(b"<t> </t>", xml)
        self.assertIn(b'<t xml:space="preserve"> </t>', xml)

    def test_expands_asset_rows_when_assignment_has_more_than_template_capacity(self):
        for orden in range(1, 8):
            self.crear_detalle(orden=orden, serie=f"SERIE{orden:03d}")

        _acta, workbook = self.cargar_workbook_generado()
        ws = workbook.active

        self.assertEqual(ws["B20"].value, "Laptop")
        self.assertIn("SERIE007", ws["H20"].value)
        self.assertIn("CREDENCIALES DE ACCESO", str(ws["B21"].value))
